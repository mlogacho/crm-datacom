"""Excel report generator for monthly billing — CONTROL FACTURAS format.

Data source: ClientService (status=INSTALLED) from the Clients/Services module.
This report is independent of the BillingRecord import module.
"""

import io
from collections import OrderedDict
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import MONTH_CHOICES

MONTH_NAMES = dict(MONTH_CHOICES)

# Statuses that represent a billable (recurring) service
BILLABLE_STATUSES = ('INSTALLED',)


def _resolve_logo():
    """Return absolute path to DataCom logo, or None if not found."""
    import os

    candidates = [
        os.path.join(os.path.dirname(__file__), '..', 'frontend', 'public', 'datacom_logo.png'),
        os.path.join(os.path.dirname(__file__), '..', 'frontend', 'dist', 'datacom_logo.png'),
        '/var/www/crm-datacom/frontend/public/datacom_logo.png',
        '/var/www/crm-datacom/frontend/dist/datacom_logo.png',
    ]
    for path in candidates:
        normalised = os.path.normpath(path)
        if os.path.isfile(normalised):
            return normalised
    return None


def get_report_data(mes, anio):
    """
    Build report data from ClientService records.

    Recurring rows  — services with agreed_price > 0 (MRC).
    Additional rows — services with nrc > 0 (NRC, one-time charges).

    Returns
    -------
    dict:
        'clients':     OrderedDict { client_pk: {name, records, total} }
        'additionals': list of {client_name, service_label, service_amount,
                                iva_amount, total, observations, factura, credito}
    """
    import calendar
    from datetime import date
    from django.db.models import Q
    from services.models import ClientService

    qs = (
        ClientService.objects
        .select_related('client', 'service')
        .filter(status__in=BILLABLE_STATUSES, client__is_active=True)
        .order_by('client__name', 'service__name', 'id')
    )

    if mes and anio:
        last_day     = calendar.monthrange(anio, mes)[1]
        period_start = date(anio, mes, 1)
        period_end   = date(anio, mes, last_day)
        qs = qs.filter(start_date__lte=period_end).filter(
            Q(end_date__isnull=True) | Q(end_date__gte=period_start)
        )

    clients_map = OrderedDict()
    additionals = []

    for cs in qs:
        # ── Recurring (MRC) ──────────────────────────────────────────────
        amount = Decimal(str(cs.agreed_price or 0))
        if amount > 0:
            key = cs.client_id
            if key not in clients_map:
                clients_map[key] = {
                    'name':    cs.client.name,
                    'records': [],
                    'total':   Decimal('0'),
                }
            iva   = round(amount * Decimal('0.15'), 2)
            total = round(amount + iva, 2)
            clients_map[key]['records'].append({
                'service_label':  cs.service.name if cs.service else '',
                'service_amount': float(amount),
                'iva_amount':     float(iva),
                'total':          float(total),
                'observations':   cs.notes or '',
                'factura':        '',
                'credito':        '',
            })
            clients_map[key]['total'] += amount

        # ── Non-recurring (NRC) — only in the month the service started ──
        nrc_amount = Decimal(str(cs.nrc or 0))
        if nrc_amount > 0:
            # NRC is a one-time charge: include only if start_date falls within the
            # selected month (or always when no month filter is applied).
            include_nrc = True
            if mes and anio:
                sd = cs.start_date
                include_nrc = (sd is not None and sd.year == anio and sd.month == mes)
            if include_nrc:
                nrc_iva   = round(nrc_amount * Decimal('0.15'), 2)
                nrc_total = round(nrc_amount + nrc_iva, 2)
                additionals.append({
                    'client_name':    cs.client.name,
                    'service_label':  cs.service.name if cs.service else '',
                    'service_amount': float(nrc_amount),
                    'iva_amount':     float(nrc_iva),
                    'total':          float(nrc_total),
                    'observations':   cs.notes or '',
                    'factura':        '',
                    'credito':        '',
                })

    return {'clients': clients_map, 'additionals': additionals}


def _thin_border():
    s = Side(style='thin', color='FF000000')
    return Border(left=s, right=s, top=s, bottom=s)


def _money_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=round(float(value or 0), 2))
    cell.number_format = '$#,##0.00'
    return cell


def generate_billing_excel(mes, anio):
    """
    Build an xlsx report matching the CONTROL FACTURAS RECURRENTES format.

    Parameters
    ----------
    mes  : int | None  — billing month (1-12) or None for full-year report
    anio : int         — billing year

    Returns
    -------
    (io.BytesIO, str)  — file buffer and month label for file naming.
    """
    mes_label   = MONTH_NAMES.get(mes, '') if mes else ''
    result      = get_report_data(mes, anio)
    clients_map = result['clients']
    additionals = result['additionals']

    # ── Workbook / sheet ──────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = mes_label if mes_label else str(anio)

    # Column widths: A    B     C     D     E     F     G     H     I
    col_widths = [29.5, 70.0, 17.5, 16.5, 16.5, 25.0, 14.5, 13.0, 13.0]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Row heights
    ws.row_dimensions[1].height = 12
    ws.row_dimensions[2].height = 42
    ws.row_dimensions[3].height = 12
    ws.row_dimensions[4].height = 26
    ws.row_dimensions[5].height = 8
    ws.row_dimensions[6].height = 32

    # ── Logo ──────────────────────────────────────────────────────────────
    logo_path = _resolve_logo()
    if logo_path:
        try:
            img = XLImage(logo_path)
            img.width  = 160
            img.height = 63
            ws.add_image(img, 'A1')
            # Second logo on the right
            img2 = XLImage(logo_path)
            img2.width  = 160
            img2.height = 63
            ws.add_image(img2, 'H1')
        except Exception:
            pass

    # ── Title row 4 ───────────────────────────────────────────────────────
    ws.merge_cells('A4:I4')
    title_text = f'FACTURACION MENSUAL RECURRENTE {anio}'
    title            = ws['A4']
    title.value      = title_text
    title.font       = Font(name='Arial', bold=True, size=12, color='001E41')
    title.alignment  = Alignment(horizontal='center', vertical='center')

    # ── Header row 6 ─────────────────────────────────────────────────────
    HDR_FILL  = PatternFill('solid', fgColor='001E41')
    HDR_FONT  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    HDR_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
    BDR       = _thin_border()

    HEADERS = [
        'Cliente', 'Servicio por Cliente', 'Servicio sin IVA',
        '15% IVA', 'TOTAL', 'Facturacion Total Clientes',
        'OBSERVACIONES', 'FACTURA', 'CREDITO',
    ]
    for col, h in enumerate(HEADERS, 1):
        cell           = ws.cell(row=6, column=col, value=h)
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = HDR_ALIGN
        cell.border    = BDR

    # ── Shared styles ─────────────────────────────────────────────────────
    RIGHT    = Alignment(horizontal='right',  vertical='center')
    LEFT     = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    CENTER   = Alignment(horizontal='center', vertical='center')
    DATA_F   = Font(name='Arial', size=11)
    BOLD_F   = Font(name='Arial', bold=True, size=11)
    CLI_FILL = PatternFill('solid', fgColor='001E41')   # Azul Datacom for client names
    CLI_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')  # white text on Azul Datacom

    # ── Data rows ─────────────────────────────────────────────────────────
    row_cursor = 7

    for client_data in clients_map.values():
        records      = client_data['records']
        client_name  = client_data['name']
        client_total = float(client_data['total'])
        start_row    = row_cursor

        for i, rec in enumerate(records):
            r = row_cursor
            ws.row_dimensions[r].height = 18

            # A: client name (first row only; merged later)
            if i == 0:
                a           = ws.cell(row=r, column=1, value=client_name)
                a.font      = CLI_FONT
                a.alignment = LEFT
                a.fill      = CLI_FILL

            # B: service label
            b           = ws.cell(row=r, column=2, value=rec['service_label'])
            b.font      = DATA_F
            b.alignment = LEFT

            # C: Servicio sin IVA
            c = _money_cell(ws, r, 3, rec['service_amount'])
            c.font = DATA_F; c.alignment = RIGHT; c.border = BDR

            # D: 15% IVA
            d = _money_cell(ws, r, 4, rec['iva_amount'])
            d.font = DATA_F; d.alignment = RIGHT; d.border = BDR

            # E: Total
            e = _money_cell(ws, r, 5, rec['total'])
            e.font = DATA_F; e.alignment = RIGHT; e.border = BDR

            # F: Facturacion Total Clientes (first row only)
            if i == 0:
                f = _money_cell(ws, r, 6, client_total)
                f.font = BOLD_F; f.alignment = RIGHT; f.border = BDR

            # G: Observaciones
            g           = ws.cell(row=r, column=7, value=rec['observations'])
            g.font      = DATA_F; g.alignment = LEFT; g.border = BDR

            # H: Factura
            h_           = ws.cell(row=r, column=8, value=rec['factura'])
            h_.font      = DATA_F; h_.alignment = CENTER; h_.border = BDR

            # I: Crédito
            i_           = ws.cell(row=r, column=9, value=rec['credito'])
            i_.font      = DATA_F; i_.alignment = CENTER; i_.border = BDR

            ws.cell(row=r, column=1).border = BDR
            ws.cell(row=r, column=2).border = BDR

            row_cursor += 1

        # Apply navy fill + white font to all client rows in column A
        for rr in range(start_row, row_cursor):
            cell = ws.cell(row=rr, column=1)
            cell.fill = CLI_FILL
            cell.font = CLI_FONT

        # Merge A and F across all service rows of this client
        if len(records) > 1:
            ws.merge_cells(f'A{start_row}:A{row_cursor - 1}')
            ws.cell(row=start_row, column=1).alignment = Alignment(
                horizontal='left', vertical='center', wrap_text=True
            )
            ws.merge_cells(f'F{start_row}:F{row_cursor - 1}')
            ws.cell(row=start_row, column=6).alignment = Alignment(
                horizontal='right', vertical='center'
            )

        # Empty separator row
        ws.row_dimensions[row_cursor].height = 6
        row_cursor += 1

    # ── Totals ────────────────────────────────────────────────────────────
    rec_sin_iva = sum(rec['service_amount'] for cd in clients_map.values() for rec in cd['records'])
    rec_iva     = sum(rec['iva_amount']     for cd in clients_map.values() for rec in cd['records'])
    rec_total   = sum(rec['total']          for cd in clients_map.values() for rec in cd['records'])

    add_sin_iva = sum(a['service_amount'] for a in additionals)
    add_iva     = sum(a['iva_amount']     for a in additionals)
    add_total   = sum(a['total']          for a in additionals)

    grand_sin_iva = rec_sin_iva + add_sin_iva
    grand_iva     = rec_iva     + add_iva
    grand_total   = rec_total   + add_total

    YELLOW  = PatternFill('solid', fgColor='FFFF00')
    LT_GREY = PatternFill('solid', fgColor='E8E8E8')
    TOTAL_F = Font(name='Arial', bold=True, size=12)
    GRAND_F = Font(name='Arial', bold=True, size=12)

    def _totals_row(row, label, s_iva, iva, total, font, fill=None):
        _fill = fill or YELLOW
        ws.row_dimensions[row].height = 22
        ws.merge_cells(f'A{row}:B{row}')
        lbl = ws.cell(row=row, column=1, value=label)
        lbl.font = font; lbl.alignment = LEFT
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.fill = _fill; cell.border = BDR; cell.font = font
        for col, val in [(3, s_iva), (4, iva), (5, total)]:
            m = _money_cell(ws, row, col, val)
            m.fill = _fill; m.border = BDR; m.alignment = RIGHT; m.font = font

    # TOTAL RECURRENTES
    _totals_row(row_cursor, 'TOTAL RECURRENTES', rec_sin_iva, rec_iva, rec_total, TOTAL_F)
    row_cursor += 1

    # ── ADICIONALES NO RECURRENTES section ──────────────────────────────
    ws.row_dimensions[row_cursor].height = 6
    row_cursor += 1

    # Grey section header
    ws.row_dimensions[row_cursor].height = 20
    ws.merge_cells(f'A{row_cursor}:I{row_cursor}')
    sec_hdr = ws.cell(row=row_cursor, column=1, value='ADICIONALES NO RECURRENTES')
    sec_hdr.fill = LT_GREY
    sec_hdr.font = Font(name='Arial', bold=True, size=11)
    sec_hdr.alignment = Alignment(horizontal='center', vertical='center')
    sec_hdr.border = BDR
    for col in range(2, 10):
        c = ws.cell(row=row_cursor, column=col)
        c.fill = LT_GREY; c.border = BDR
    row_cursor += 1

    # Additional rows
    for add in additionals:
        r = row_cursor
        ws.row_dimensions[r].height = 18
        a_col = ws.cell(row=r, column=1, value=add['client_name'])
        a_col.font = BOLD_F; a_col.alignment = LEFT; a_col.border = BDR
        b_col = ws.cell(row=r, column=2, value=add['service_label'])
        b_col.font = DATA_F; b_col.alignment = LEFT; b_col.border = BDR
        c_col = _money_cell(ws, r, 3, add['service_amount'])
        c_col.font = DATA_F; c_col.alignment = RIGHT; c_col.border = BDR
        d_col = _money_cell(ws, r, 4, add['iva_amount'])
        d_col.font = DATA_F; d_col.alignment = RIGHT; d_col.border = BDR
        e_col = _money_cell(ws, r, 5, add['total'])
        e_col.font = DATA_F; e_col.alignment = RIGHT; e_col.border = BDR
        f_col = ws.cell(row=r, column=6, value='')
        f_col.border = BDR
        g_col = ws.cell(row=r, column=7, value=add['observations'])
        g_col.font = DATA_F; g_col.alignment = LEFT; g_col.border = BDR
        h_col = ws.cell(row=r, column=8, value=add['factura'])
        h_col.font = DATA_F; h_col.alignment = CENTER; h_col.border = BDR
        i_col = ws.cell(row=r, column=9, value=add['credito'])
        i_col.font = DATA_F; i_col.alignment = CENTER; i_col.border = BDR
        row_cursor += 1

    # TOTAL ADICIONALES
    ws.row_dimensions[row_cursor].height = 6
    row_cursor += 1
    _totals_row(row_cursor, 'TOTAL ADICIONALES', add_sin_iva, add_iva, add_total, TOTAL_F)
    row_cursor += 2

    # TOTAL FACTURACION (RECURRENTES + ADICIONALES)
    _totals_row(row_cursor, 'TOTAL FACTURACION', grand_sin_iva, grand_iva, grand_total, GRAND_F)
    row_cursor += 3

    # ── Summary: per-client totals ────────────────────────────────────────
    ws.merge_cells(f'A{row_cursor}:B{row_cursor}')
    sh = ws.cell(row=row_cursor, column=1, value=f'FACTURACION MENSUAL RECURRENTE {anio}')
    sh.font = Font(name='Arial', bold=True, size=12, color='001E41')
    sh.alignment = Alignment(horizontal='center', vertical='center')
    row_cursor += 1

    for cd in clients_map.values():
        r = row_cursor
        ws.cell(row=r, column=1, value=cd['name']).font = BOLD_F
        m = _money_cell(ws, r, 2, cd['total'])
        m.font = BOLD_F
        row_cursor += 1

    if additionals:
        r = row_cursor
        ws.cell(row=r, column=1, value='TOTAL ADICIONALES').font = BOLD_F
        m = _money_cell(ws, r, 2, add_sin_iva)
        m.font = BOLD_F
        row_cursor += 1

    # ── Print settings ────────────────────────────────────────────────────
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, mes_label
